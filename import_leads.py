#!/usr/bin/env python3
"""
import_leads.py — Import xlsx lead files into the leads table.

Non-customer leads  → status preserved from source (status_name column)
Previous customers  → status forced to 'sold'

Usage:
    python import_leads.py
"""

import openpyxl
import click
from db import get_conn, init_db

FILES = [
    ("data/MKE_MAD_NonCustomer_Leads.xlsx",    False),
    ("data/MN_NonCustomer_Leads.xlsx",          False),
    ("data/MKE_MAD_Previous_Customers.xlsx",    True),
    ("data/MN_Previous_Customers.xlsx",         True),
]

CREATE_TABLE = """
CREATE TABLE IF NOT EXISTS leads (
    id               SERIAL PRIMARY KEY,
    source_id        TEXT,
    user_id          TEXT,
    lead_owner       TEXT,
    lead_active      TEXT,
    business_name    TEXT,
    import_batch_id  TEXT,
    first_name       TEXT,
    last_name        TEXT,
    email            TEXT,
    phone_primary    TEXT,
    phone_secondary  TEXT,
    appointment_time TEXT,
    street1          TEXT,
    street2          TEXT,
    city             TEXT,
    state            TEXT,
    postal_code      TEXT,
    latitude         DOUBLE PRECISION,
    longitude        DOUBLE PRECISION,
    status_id        INTEGER REFERENCES lead_statuses(id),
    note             TEXT,
    form_data        TEXT,
    updated_at       TEXT,
    inserted_at      TEXT,
    deleted          TEXT,
    source_file_id   INTEGER REFERENCES source_files(id),
    test_lead        INTEGER NOT NULL DEFAULT 0
)
"""

# Maps xlsx column name → table column name
COL_MAP = {
    "id":                           "source_id",
    "user_id":                      "user_id",
    "lead_owner":                   "lead_owner",
    "lead_active":                  "lead_active",
    "business_name":                "business_name",
    "import_batch_id":              "import_batch_id",
    "contacts_first_name":          "first_name",
    "contacts_last_name":           "last_name",
    "contacts_email":               "email",
    "contacts_phone_primary":       "phone_primary",
    "contacts_phone_secondary":     "phone_secondary",
    "contacts_appointment_time":    "appointment_time",
    "address_street1":              "street1",
    "address_street2":              "street2",
    "address_city":                 "city",
    "address_state":                "state",
    "address_postal_code":          "postal_code",
    "address_latitude":             "latitude",
    "address_longitude":            "longitude",
    "status_name":                  "status",
    "note":                         "note",
    "form_data":                    "form_data",
    "updated_at":                   "updated_at",
    "inserted_at":                  "inserted_at",
    "deleted":                      "deleted",
}


def _get_or_create(conn, table: str, col: str, value) -> int | None:
    """Return the id for value in a lookup table, inserting a new row if needed."""
    if value is None:
        return None
    conn.execute(
        f"INSERT INTO {table} ({col}) VALUES (%s) ON CONFLICT ({col}) DO NOTHING",
        (value,)
    )
    return conn.execute(
        f"SELECT id FROM {table} WHERE {col} = %s", (value,)
    ).fetchone()["id"]


@click.command()
@click.option("--reset", is_flag=True, help="Drop and recreate the leads table first")
def import_leads(reset: bool):
    """Import all xlsx lead files into the leads table."""
    init_db()
    conn = get_conn()

    if reset:
        conn.execute("DROP TABLE IF EXISTS leads")
        click.secho("  Dropped existing leads table", fg="yellow")

    conn.execute(CREATE_TABLE)
    conn.commit()

    total = 0
    for filepath, is_customer in FILES:
        # Ensure the source file has a lookup row and grab its id once per file
        source_file_id = _get_or_create(conn, "source_files", "path", filepath)
        conn.commit()

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        rows_inserted = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))

            mapped = {db_col: record.get(xlsx_col) for xlsx_col, db_col in COL_MAP.items()}
            if is_customer:
                mapped["status"] = "sold"

            # Convert text lookups → FK ids
            mapped["status_id"]      = _get_or_create(conn, "lead_statuses", "name", mapped.pop("status", None))
            mapped["source_file_id"] = source_file_id

            cols   = list(mapped.keys())
            values = list(mapped.values())
            placeholders = ", ".join(["%s"] * len(cols))
            col_names    = ", ".join(cols)

            conn.execute(
                f"INSERT INTO leads ({col_names}) VALUES ({placeholders})",
                values
            )
            rows_inserted += 1

        conn.commit()
        wb.close()
        label = "customers" if is_customer else "leads"
        click.echo(f"  {filepath}: {rows_inserted:,} {label} imported")
        total += rows_inserted

    click.secho(f"\n✅  {total:,} total rows imported into leads table\n", fg="green")
    conn.close()


if __name__ == "__main__":
    import_leads()
